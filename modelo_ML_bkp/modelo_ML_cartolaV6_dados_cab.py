import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

ARQUIVOS_CSV = {
    'Atletas mercado': 'atletas_mercado.csv',
    'Dados Destaque': 'dados_destaque.csv',
    'Dados Partidas': 'dados_partidas.csv',
    'Dados Times': 'dados_times.csv',
    'Pontuacoes Jogadores': 'pontuacoes_jogadores.csv'
}

SCOUT_WEIGHTS = {
    'DS': 1.2,
    'FC': -0.3,
    'GC': -3.0,
    'CA': -1.0,
    'CV': -3.0,
    'SG': 5.0,
    'DE': 1.0,
    'DP': 7.0,
    'GS': -1.0,
    'PC': -1.0,
    'FS': 0.5,
    'I': -0.1,
    'A': 5.0,
    'FT': 3.0,
    'FD': 1.2,
    'FF': 0.8,
    'G': 8.0,
    'PP': -4.0,
    'PS': 1.0,
    'V': 1.0
}

def carregar_dados():
    diretorio = 'bd_cartola_modelo'
    tabelas = {}

    for nome_tabela, nome_arquivo in ARQUIVOS_CSV.items():
        caminho_arquivo = os.path.join(diretorio, nome_arquivo)
        if os.path.exists(caminho_arquivo):
            tabela = pd.read_csv(caminho_arquivo)
            tabelas[nome_tabela] = tabela
        else:
            print(f'Arquivo não encontrado: {caminho_arquivo}')

    if len(tabelas) != len(ARQUIVOS_CSV):
        # Verificar quais arquivos estão faltando
        arquivos_faltantes = set(ARQUIVOS_CSV.values()) - set(tabelas.keys())
        for arquivo in arquivos_faltantes:
            print(f'Arquivo faltando: {arquivo}')

    return tabelas


def preencher_campos_vazios(tabelas):
    for nome_tabela, tabela in tabelas.items():
        tabela.fillna(0, inplace=True)

    # Tratamento específico para a tabela "Atletas mercado"
    if 'Atletas mercado' in tabelas:
        tabela_atletas = tabelas['Atletas mercado']
        tabela_atletas.fillna(0, inplace=True)
        tabela_atletas[['minimo_para_valorizar']] = tabela_atletas[['minimo_para_valorizar']].astype(int)
        for scout in SCOUT_WEIGHTS.keys():
            if scout in tabela_atletas:
                tabela_atletas[scout] = tabela_atletas[scout] * SCOUT_WEIGHTS[scout]

    # Tratamento específico para a tabela "Pontuacoes Jogadores"
    if 'Pontuacoes Jogadores' in tabelas:
        tabela_pontuacoes = tabelas['Pontuacoes Jogadores']
        tabela_pontuacoes.fillna(0, inplace=True)
        for scout in SCOUT_WEIGHTS.keys():
            if scout in tabela_pontuacoes:
                tabela_pontuacoes[scout] = tabela_pontuacoes[scout] * SCOUT_WEIGHTS[scout]


def criar_dataframe(tabelas):
    # Selecionar apenas as tabelas relevantes para o modelo
    tabelas_selecionadas = ['Atletas mercado', 'Dados Partidas', 'Dados Times', 'Pontuacoes Jogadores']
    tabelas_filtradas = {nome_tabela: tabela for nome_tabela, tabela in tabelas.items() if nome_tabela in tabelas_selecionadas and nome_tabela != 'Dados Destaque'}


    # Criar DataFrame vazio
    df = pd.DataFrame()

    # Mesclar as tabelas relevantes no DataFrame
    for nome_tabela, tabela in tabelas_filtradas.items():
        tabela.columns = pd.MultiIndex.from_product([[nome_tabela], tabela.columns])
        df = pd.concat([df, tabela], axis=1)

    # Preencher campos vazios com zero
    df.fillna(0, inplace=True)

    # Aplicar pesos aos scouts na tabela "Pontuacoes Jogadores"
    if 'Pontuacoes Jogadores' in tabelas_filtradas:
        tabela_pontuacoes = df['Pontuacoes Jogadores']
        for scout, peso in SCOUT_WEIGHTS.items():
            if scout in tabela_pontuacoes:
                tabela_pontuacoes[scout] = tabela_pontuacoes[scout] * peso

    # Renomear campos
    df.rename(columns={'clube_casa_nome': 'clube_nome', 'clube_visitante_nome': 'clube_nome', 'clube_casa_posicao': 'clube_posicao', 'clube_visitante_posicao': 'clube_posicao', 'aproveitamento_mandante': 'clube_aproveitamento', 'aproveitamento_visitante': 'clube_aproveitamento'}, level=0, inplace=True)

    # Criar campo "clube_mand" indicando se o time do atleta é mandante ou não
    df['clube_mand'] = df['Pontuacoes Jogadores']['clube_id'] == df['Dados Partidas']['clube_casa_id']

    return df



# Passo 1: Carregar os dados dos arquivos CSV
tabelas = carregar_dados()

# Passo 2: Preencher campos vazios com zero
preencher_campos_vazios(tabelas)

# Passo 3: Criar DataFrame único com as tabelas relevantes
df = criar_dataframe(tabelas)

# Verificar se todas as tabelas foram carregadas corretamente após o preenchimento
tabelas_completas = True
for nome_tabela, tabela in tabelas.items():
    if tabela.isna().any().any():
        print(f'Valores ausentes encontrados na tabela: {nome_tabela}')
        colunas_ausentes = tabela.columns[tabela.isna().any()].tolist()
        print(f'Colunas com valores ausentes: {colunas_ausentes}')
        tabelas_completas = False

if tabelas_completas:
    print('Limpeza dos dados concluída. Os arquivos CSV limpos foram salvos.')
    print('DataFrame consolidado:')
    print(df.head())

    # Criar arquivo do Excel com o DataFrame consolidado e a tabela de origem
    nome_arquivo_excel = f'dados_consolidados_{pd.Timestamp.now().strftime("%Y%m%d%H%M%S")}.xlsx'
    workbook = Workbook()

    # Exportar DataFrame consolidado
    df_sheet = workbook.create_sheet('Dados Consolidados')
    df_header = [column[1] for column in df.columns]
    df_sheet.append(df_header)
    for row in dataframe_to_rows(df, index=False, header=False):
        df_sheet.append(row)

    # Exportar tabela com campos/tabelas de origem
    tabela_origem = pd.DataFrame({'Campo': df.columns.get_level_values(0), 'Tabela': df.columns.get_level_values(1)})
    tabela_origem_sheet = workbook.create_sheet('Origem dos Dados')
    for row in dataframe_to_rows(tabela_origem, index=False, header=True):
        tabela_origem_sheet.append(row)

    # Remover a planilha vazia padrão
    workbook.remove(workbook['Sheet'])

    # Salvar o arquivo do Excel
    workbook.save(nome_arquivo_excel)

    print(f'Dados exportados para o arquivo: {nome_arquivo_excel}')
else:
    print('Erro ao carregar tabelas. Verifique se todos os arquivos estão presentes ou se há valores ausentes nos dados.')
